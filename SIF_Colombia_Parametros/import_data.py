"""
Carga TODAS las dimensiones compartidas de poc_colombia desde
"Informes FIC - Parámetros.accdb" (fuente autoritativa confirmada: sus
conteos de fila calzan exacto con las copias que ya cargaban 351/Info
Portafolio/PowerBI) mas 4 hojas que solo existen en
Bases_Colombia\\Input\\*.xlsx (no tienen tabla Access equivalente todavia).

Este proyecto REEMPLAZA a SIF_Colombia_PowerBI/import_dimensiones.py como
cargador autoritativo de tbl_Fondos/tbl_Arrendatarios (ver ese directorio,
marcado .legacy). Tambien reemplaza la carga de dimensiones que hacian
SIF_Colombia_351 (tbl_Fechas, tbl_Inmuebles, tbl_Contratos, tbl_Usuarios,
VL_Disp_xa_Venta) e SIF_Colombia_Info_portafolio (tbl_Niveles, tbl_Tiempos,
tbl_Cuentas, tbl_Totales, tbl_Centros) -- esos dos proyectos quedaron
recortados a solo sus hechos/EEFF propios.

Flujo igual al resto de los proyectos: staging (stg.<tabla>, TRUNCATE +
insert masivo) + etl.usp_ReemplazarDimension (ver
poc_colombia/schema/09_sp_etl.sql). Ninguna tabla de este script es un
hecho particionado, por eso todo usa usp_ReemplazarDimension.

tbl_Contratos_Detalle es especial: se arma por UNION DE NOMBRE DE COLUMNA
entre la tabla Access "Contratos" (detalle de contratos VIGENTE, 78
columnas) y la hoja "Contratos no vigentes" de Contratos.xlsx (68 columnas,
con renombres/columnas propias) -- ver poc_colombia/schema/11_detalle_
contratos_inmuebles.sql para el detalle de la union de columnas.

Pensado para correrse en tu propia terminal (no via un agente), porque pide
la contrasena de SQL de forma interactiva con getpass.

USO:
    1) Revisar el esquema real de Access antes de importar (no toca SQL Server):
       python import_data.py --check-schema-only

    2) Import real (re-corrible sin duplicar):
       python import_data.py --sql-server "mi_servidor" --sql-user "mi_usuario"

Parametros opcionales:
    --access-path   Ruta al .accdb (por defecto, la ruta ya conocida)
    --input-dir     Ruta a Bases_Colombia\\Input (por defecto, la ya conocida)
    --sql-database  Nombre de la base destino (por defecto poc_colombia)
"""

import argparse
import datetime
import decimal
import getpass
import os
import sys

import openpyxl
import pyodbc

DEFAULT_ACCESS_PATH = (
    r"C:\Users\pal2101\OneDrive - Patria Investimentos\Producción\Utilidades"
    r"\Colombia\Bases_Colombia\Informes FIC - Parámetros.accdb"
)
DEFAULT_INPUT_DIR = (
    r"C:\Users\pal2101\OneDrive - Patria Investimentos\Producción\Utilidades"
    r"\Colombia\Bases_Colombia\Input"
)

# Tablas 1:1 (mismo nombre en Access y en dbo/stg), todas "dimension"
# (usp_ReemplazarDimension) -- ninguna de estas es un hecho particionado.
SIMPLE_ACCESS_TABLES = [
    "tbl_Fechas", "tbl_Niveles", "tbl_Tiempos", "tbl_Totales", "tbl_Cuentas",
    "tbl_Fondos", "tbl_Arrendatarios", "tbl_Centros", "tbl_Usuarios",
    "tbl_Inmuebles", "tbl_Contratos", "VL_Disp_xa_Venta",
    "tbl_Inmuebles_Otros", "tbl_MI",
]

# (tabla_origen_Access, tabla_destino) -- nombre distinto entre origen y destino.
RENAMED_ACCESS_TABLES = [
    ("CONTRATOS_OTROS", "tbl_Contratos_Otros_Fondos"),
    ("CONTRATOS_RESTITUIDOS", "tbl_Contratos_Restituidos_OtrosFondos"),
]

# Columnas de origen con tilde que en dbo/stg se guardaron sin tilde
# (consistencia con el resto del modelo, ej. tbl_Centros.Tipologia).
COLUMN_ALIASES = {
    "tbl_Inmuebles_Otros": {"Tipología": "Tipologia"},
    "tbl_Contratos_Otros_Fondos": {"Tipología": "Tipologia"},
    "tbl_Contratos_Restituidos_OtrosFondos": {"Tipología": "Tipologia"},
}

# Mismo placeholder que ya vivia en SIF_Colombia_PowerBI/import_dimensiones.py
# (ahora legacy) -- Parametros.accdb tampoco trae una fila COD_FONDO=0, sigue
# haciendo falta como centinela "no aplica a un fondo especifico".
FONDO_PLACEHOLDER = (0, None, "N/A", "No aplica", "No aplica / Consolidado (placeholder)", None, None)

# Columnas de dbo/stg.tbl_Contratos_Detalle, en el mismo orden que
# poc_colombia/schema/11_detalle_contratos_inmuebles.sql (sin Id).
CONTRATOS_DETALLE_COLUMNS = [
    "Origen_Hoja", "FECHA", "COD", "COD_Nuevo", "ARRENDATARIO", "CIUDAD DEL PREDIO",
    "TIPOLOGÍA INMUEBLE", "DESCRIPCIÓN", "TIPO DE CONTRATO",
    "ÁREA RENTABLE CONTRATO ARRENDAMIENTO", "ÚLTIMO CANON", "CANON M2",
    "FECHA INICIO CONTRATO", "FECHA INICIO COBRO CANON", "DURACIÓN AÑOS ARRENDAMIENTO",
    "FECHA FINALIZACIÓN CONTRATO INICIAL",
    "FECHA FIN CONTRATO (incluye prórrogas/renovaciones)", "MES DE AJUSTE ANUAL",
    "INCREMENTO ANUAL SOBRE IPC", "INCREMENTO ANUAL COMO VECES IPC", "DETALLE INCREMENTO",
    "CANON VARIABLE", "CLÁUSULA INCREMENTO DE CANON", "PREAVISO EN MESES TÉRMINO INICIAL",
    "PREAVISO PRÓRROGA / RENOVACIÓN", "PRÓRROGAS / RENOVACIONES",
    "CLÁUSULA CONDICIONES PRÓRROGAS / RENOVACIONES", "AÑOS POR PRÓRROGA/RENOVACIÓN",
    "NÚMERO DE PRÓRRGAS / RENOVACIONES", "FECHA PRÓRROGA / RENOVACIÓN", "FECHA PREAVISO",
    "CLÁUSULA SALIDA ANTICIPADA", "FECHA POSIBLE DE TERMINACIÓN ANTICIPADA",
    "CLÁUSULA REVISIÓN CONDICIONES ECONÓMICAS", "PISO / TECHO RENEGOCIACIÓN CANON",
    "REVISIÓN CANON TÉRMINO INICIAL (AÑOS)", "REVISIÓN CANON PRORROGA / RENOVACIÓN (AÑOS)",
    "PRÓXIMA FECHA REVISIÓN DE CONDICIONES ECONÓMICAS",
    "MESES PARA REVISIÓN DE CONDICIONES ECONÓMICAS",
    "ACTIVACIÓN FECHA REVISIÓN DE CONDICIONES ECONÓMICAS", "CONDICIONES ESPECIALES DE CANON",
    "CONDICIONES DE FACTURACIÓN", "DIA DE INCREMENTO", "DIAS DE VENCIMIENTO DE FACTURA",
    "DIAS DE VENCIMIENTO DE FACTURA (HÁBIL O CALENDARIO)", "BASE DE FACTURACIÓN",
    "PERIODICIDAD DE LA FACTURACION", "FORMULA DE INCREMENTO", "TIPO DE IPC INCREMENTO",
    "CANON MOBILIARIO (SI APLICA)", "CANON AVISO (SI APLICA)", "REPARACIONES NECESARIAS",
    "REPARACIONES LOCATIVAS", "MANTENIMIENTO DEL INMUEBLE", "REPOSICIONES MUEBLES Y EQUIPOS",
    "CUOTAS DE ADMINISTRACIÓN ORDINARIAS", "CUOTAS DE ADMINISTRACIÓN EXTRAORDINARIAS",
    "SERVICIOS PÚBLICOS", "SEGUROS INMUEBLE", "GARANTÍAS", "DETALLE GARANTÍAS", "COVENANTS",
    "OTROSÍES", "FECHA OTROSÍ", "TIPO DE PERSONA", "NIT ARRENDATARIO", "COD INMUEBLE",
    "COD_INMUEBLE_Nuevo", "INMUEBLE CONSOLIDADO", "ENTIDAD ESTATAL", "SE FACTURA CON IVA O NO",
    "OBSERVACIONES", "ESTANDARIZAR ARCHIVOS DIGITALES", "MOBILIARIO Y OBRA BLANCA",
    "CANON DE MERCADO", "SECTOR ECONÓMICO", "VALIDACIÓN ÁREA DE CONTRATO ARRENDAMIENTO",
    "DESCRIPCIÓN ANTIGUA", "DIRECCIÓN", "OBSERVACIONES REVISIÓN DE CONTRATO",
]

# La hoja "Contratos no vigentes" usa 'CODIGO FRO' en la posicion de COD.
CONTRATOS_NO_VIGENTES_ALIASES = {"CODIGO FRO": "COD"}

# Hoja "Inm" de Inmuebles.xlsx -> tbl_Inmuebles_Estrategia (acentos del origen
# normalizados a los nombres de columna sin tilde ya usados en el resto del modelo).
INM_ESTRATEGIA_ALIASES = {
    "Ubicación": "Ubicacion", "Dirección": "Direccion", "Tipología": "Tipologia",
    "Certificación": "Certificacion", "Subtipología": "Subtipologia",
    "Destinación": "Destinacion", "Operación": "Operacion", "Acción": "Accion",
}

# Hoja "MI Contratos" de Inmuebles.xlsx -> tbl_Inmuebles_MI_Contratos
# ("ARREDATARIO" es un typo real del origen, se mapea igual a Arrendatario).
MI_CONTRATOS_ALIASES = {
    "COD_INM": "Cod_Inm", "CODIGO CONTRATO": "Cod_Contrato", "ARREDATARIO": "Arrendatario",
    "Matricula": "Matricula", "Descripción": "Descripcion",
}


def access_connection(access_path):
    conn_str = r"Driver={Microsoft Access Driver (*.mdb, *.accdb)};" f"DBQ={access_path};"
    return pyodbc.connect(conn_str, autocommit=True)


def sql_connection(server, database, user, password):
    conn_str = (
        "Driver={ODBC Driver 17 for SQL Server};"
        f"Server={server};Database={database};"
        f"UID={user};PWD={password};"
        "Encrypt=yes;TrustServerCertificate=yes;"
    )
    return pyodbc.connect(conn_str, autocommit=False)


def _clean_row(row):
    # Access CURRENCY llega via pyodbc como decimal.Decimal; fast_executemany
    # tiene un bug conocido con Decimal sin importar el tipo de columna
    # destino -- convertir a float lo evita (ya visto en 351/Info Portafolio).
    return tuple(float(v) if isinstance(v, decimal.Decimal) else v for v in row)


def _norm(name):
    return name.strip() if isinstance(name, str) else name


def reemplazar_dimension(sql_conn, table_name):
    sql_conn.cursor().execute(
        "EXEC etl.usp_ReemplazarDimension @TablaDestino = ?, @TablaStaging = ?",
        f"dbo.{table_name}", f"stg.{table_name}",
    )


def load_simple_access_table(access_conn, sql_conn, source_table, dest_table, aliases=None):
    aliases = aliases or {}
    cursor = access_conn.cursor()
    cursor.execute(f"SELECT * FROM [{source_table}]")
    cols = [aliases.get(_norm(d[0]), _norm(d[0])) for d in cursor.description]
    rows = [_clean_row(r) for r in cursor.fetchall()]

    if dest_table == "tbl_Fondos":
        # El centinela COD_FONDO=0 debe quedar en el MISMO delete+insert que
        # hace usp_ReemplazarDimension -- si se agrega despues (fuera de esa
        # transaccion), la revalidacion de FK (tbl_Centros.Cod_Fondo, etc.)
        # falla mientras el 0 no existe todavia (visto en la primera corrida).
        rows = rows + [FONDO_PLACEHOLDER]

    sql_cursor = sql_conn.cursor()
    sql_cursor.execute(f"TRUNCATE TABLE stg.[{dest_table}]")
    col_list = ", ".join(f"[{c}]" for c in cols)
    placeholders = ", ".join("?" for _ in cols)
    sql_cursor.fast_executemany = True
    sql_cursor.executemany(f"INSERT INTO stg.[{dest_table}] ({col_list}) VALUES ({placeholders})", rows)

    reemplazar_dimension(sql_conn, dest_table)
    sql_conn.commit()
    return len(rows)


def _insert_rows_by_column(sql_conn, dest_table, columns, rows):
    sql_cursor = sql_conn.cursor()
    sql_cursor.execute(f"TRUNCATE TABLE stg.[{dest_table}]")
    col_list = ", ".join(f"[{c}]" for c in columns)
    placeholders = ", ".join("?" for _ in columns)
    # SIN fast_executemany: pyodbc no lo soporta de forma confiable con
    # columnas NVARCHAR(MAX) (tbl_Contratos_Detalle tiene varias) --
    # "Restricted data type attribute violation" visto en la primera corrida.
    # Volumen de estas tablas (<10K filas) hace la diferencia de performance
    # irrelevante.
    sql_cursor.executemany(f"INSERT INTO stg.[{dest_table}] ({col_list}) VALUES ({placeholders})", rows)
    reemplazar_dimension(sql_conn, dest_table)
    sql_conn.commit()
    return len(rows)


def _excel_rows_as_dicts(path, sheet_name, aliases=None):
    aliases = aliases or {}
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[sheet_name]
        header = [_norm(c.value) for c in next(ws.iter_rows(min_row=1, max_row=1))]
        header = [aliases.get(h, h) if h else h for h in header]
        out = []
        for values in ws.iter_rows(min_row=2, values_only=True):
            row = {}
            for name, value in zip(header, values):
                if name:
                    row[name] = value
            out.append(row)
        return out
    finally:
        wb.close()


# Unicas columnas DATETIME2 reales del stg (el resto de campos con "FECHA"
# en el nombre son NVARCHAR/FLOAT -- ver 11_detalle_contratos_inmuebles.sql).
# La hoja "Contratos no vigentes" a veces trae un numero suelto sin formato
# de fecha en estas columnas (dato mal cargado en el Excel origen, no
# reconocido por openpyxl como datetime) -- se descarta a NULL en vez de
# fallar el bind de pyodbc ("Operand type clash: int is incompatible with
# datetime2").
CONTRATOS_DETALLE_DATE_COLUMNS = {
    "FECHA", "FECHA INICIO CONTRATO", "FECHA INICIO COBRO CANON",
    "FECHA FINALIZACIÓN CONTRATO INICIAL",
    "FECHA FIN CONTRATO (incluye prórrogas/renovaciones)",
}

# Columnas FLOAT del stg (numericas en Access/Excel) -- la hoja "Contratos no
# vigentes" a veces trae texto ("NA", etc.) en estas celdas; se descarta a
# NULL en vez de fallar el bind ("Error converting data type nvarchar to float").
CONTRATOS_DETALLE_FLOAT_COLUMNS = {
    "COD_Nuevo", "ÁREA RENTABLE CONTRATO ARRENDAMIENTO", "ÚLTIMO CANON", "CANON M2",
    "DURACIÓN AÑOS ARRENDAMIENTO", "MES DE AJUSTE ANUAL", "INCREMENTO ANUAL COMO VECES IPC",
    "PREAVISO EN MESES TÉRMINO INICIAL", "PREAVISO PRÓRROGA / RENOVACIÓN",
    "AÑOS POR PRÓRROGA/RENOVACIÓN", "FECHA PREAVISO", "REVISIÓN CANON TÉRMINO INICIAL (AÑOS)",
    "REVISIÓN CANON PRORROGA / RENOVACIÓN (AÑOS)", "PRÓXIMA FECHA REVISIÓN DE CONDICIONES ECONÓMICAS",
    "MESES PARA REVISIÓN DE CONDICIONES ECONÓMICAS", "ACTIVACIÓN FECHA REVISIÓN DE CONDICIONES ECONÓMICAS",
    "OTROSÍES", "FECHA OTROSÍ", "COD INMUEBLE", "COD_INMUEBLE_Nuevo",
}


def load_contratos_detalle(access_conn, sql_conn, contratos_xlsx_path):
    rows = []

    cursor = access_conn.cursor()
    cursor.execute("SELECT * FROM [Contratos]")
    cols = [_norm(d[0]) for d in cursor.description]
    for r in cursor.fetchall():
        d = {"Origen_Hoja": "VIGENTE"}
        d.update(zip(cols, _clean_row(r)))
        rows.append(d)

    for d in _excel_rows_as_dicts(contratos_xlsx_path, "Contratos no vigentes", CONTRATOS_NO_VIGENTES_ALIASES):
        d["Origen_Hoja"] = "NO VIGENTE"
        rows.append(d)

    def _value(d, c):
        v = d.get(c)
        if c in CONTRATOS_DETALLE_DATE_COLUMNS and not isinstance(v, datetime.datetime):
            return None
        if c in CONTRATOS_DETALLE_FLOAT_COLUMNS and not isinstance(v, (int, float)):
            return None
        return v

    tuples = [tuple(_value(d, c) for c in CONTRATOS_DETALLE_COLUMNS) for d in rows]
    return _insert_rows_by_column(sql_conn, "tbl_Contratos_Detalle", CONTRATOS_DETALLE_COLUMNS, tuples)


def load_inmuebles_estrategia(sql_conn, inmuebles_xlsx_path):
    columns = [
        "Cod_Inm", "Nom_Inm", "Desc_Inm", "Ubicacion", "Ubic_Pol_Inv", "Direccion", "Tipologia",
        "GLA", "Estado", "Origen", "FechaVenta", "ValorVenta", "Georef", "Coordenadas",
        "Latitud", "Longitud", "En_PA", "Certificacion", "Estrategia", "Subtipologia",
        "Destinacion", "Operacion", "PlanCortoPlazo", "Riesgo", "Accion",
    ]
    dicts = _excel_rows_as_dicts(inmuebles_xlsx_path, "Inm", INM_ESTRATEGIA_ALIASES)
    tuples = [tuple(d.get(c) for c in columns) for d in dicts if d.get("Cod_Inm") is not None]
    return _insert_rows_by_column(sql_conn, "tbl_Inmuebles_Estrategia", columns, tuples)


def load_inmuebles_compras(sql_conn, inmuebles_xlsx_path):
    # Hoja "Compras" repite encabezados (2 bloques: codificacion vieja y
    # nueva) -- openpyxl no puede mapear por nombre (colisiona), se lee por
    # POSICION: idx0=Cod_Inm (viejo, texto), idx5=Cod_Inm_Nuevo, idx6=Nom_Inm,
    # idx7=Tipologia, idx8=FechaCompra, idx9=ValorCompra (bloque nuevo, mas
    # preciso que el viejo en idx1-3).
    columns = ["Cod_Inm", "Cod_Inm_Nuevo", "Nom_Inm", "Tipologia", "FechaCompra", "ValorCompra"]
    wb = openpyxl.load_workbook(inmuebles_xlsx_path, read_only=True, data_only=True)
    try:
        ws = wb["Compras"]
        tuples = []
        for values in ws.iter_rows(min_row=2, values_only=True):
            if values[0] is None:
                continue
            cod_inm = str(values[0])
            cod_inm_nuevo = values[5] if isinstance(values[5], (int, float)) else None
            nom_inm = values[6] if isinstance(values[6], str) else None
            tipologia = values[7] if isinstance(values[7], str) else None
            fecha_compra = values[8] if isinstance(values[8], datetime.datetime) else None
            valor_compra = float(values[9]) if isinstance(values[9], (int, float, decimal.Decimal)) else None
            tuples.append((cod_inm, cod_inm_nuevo, nom_inm, tipologia, fecha_compra, valor_compra))
        return _insert_rows_by_column(sql_conn, "tbl_Inmuebles_Compras", columns, tuples)
    finally:
        wb.close()


def load_inmuebles_mi_contratos(sql_conn, inmuebles_xlsx_path):
    columns = ["Cod_Inm", "Cod_Contrato", "Arrendatario", "Matricula", "Descripcion"]
    dicts = _excel_rows_as_dicts(inmuebles_xlsx_path, "MI Contratos", MI_CONTRATOS_ALIASES)
    tuples = [tuple(d.get(c) for c in columns) for d in dicts if d.get("Cod_Inm") is not None]
    return _insert_rows_by_column(sql_conn, "tbl_Inmuebles_MI_Contratos", columns, tuples)


def check_schema(access_path, input_dir):
    access_conn = access_connection(access_path)
    try:
        cursor = access_conn.cursor()
        for table_name in SIMPLE_ACCESS_TABLES + ["Contratos"] + [t for t, _ in RENAMED_ACCESS_TABLES]:
            cursor.execute(f"SELECT * FROM [{table_name}]")
            cols = [d[0] for d in cursor.description]
            cursor.execute(f"SELECT COUNT(*) FROM [{table_name}]")
            row_count = cursor.fetchone()[0]
            print(f"== {table_name}  ({row_count} filas en Access) ==")
            print("  " + ", ".join(cols))
    finally:
        access_conn.close()

    contratos_xlsx = os.path.join(input_dir, "Contratos.xlsx")
    inmuebles_xlsx = os.path.join(input_dir, "Inmuebles.xlsx")
    for path, sheet in [(contratos_xlsx, "Contratos no vigentes"), (inmuebles_xlsx, "Inm"),
                         (inmuebles_xlsx, "Compras"), (inmuebles_xlsx, "MI Contratos")]:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb[sheet]
        header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        print(f"== {os.path.basename(path)} :: {sheet}  ({ws.max_row - 1} filas) ==")
        print("  " + ", ".join(str(h) for h in header))
        wb.close()


def import_data(access_path, input_dir, server, database, user, password):
    access_conn = access_connection(access_path)
    sql_conn = sql_connection(server, database, user, password)
    contratos_xlsx = os.path.join(input_dir, "Contratos.xlsx")
    inmuebles_xlsx = os.path.join(input_dir, "Inmuebles.xlsx")

    try:
        for table_name in SIMPLE_ACCESS_TABLES:
            print(f"Cargando {table_name}...", end=" ")
            sys.stdout.flush()
            try:
                n = load_simple_access_table(
                    access_conn, sql_conn, table_name, table_name, COLUMN_ALIASES.get(table_name)
                )
                print(f"OK ({n} filas)")
            except Exception as ex:
                sql_conn.rollback()
                print(f"ERROR: {ex}")

        for source_table, dest_table in RENAMED_ACCESS_TABLES:
            print(f"Cargando {dest_table} (desde {source_table})...", end=" ")
            sys.stdout.flush()
            try:
                n = load_simple_access_table(
                    access_conn, sql_conn, source_table, dest_table, COLUMN_ALIASES.get(dest_table)
                )
                print(f"OK ({n} filas)")
            except Exception as ex:
                sql_conn.rollback()
                print(f"ERROR: {ex}")

        for label, fn in [
            ("tbl_Contratos_Detalle", lambda: load_contratos_detalle(access_conn, sql_conn, contratos_xlsx)),
            ("tbl_Inmuebles_Estrategia", lambda: load_inmuebles_estrategia(sql_conn, inmuebles_xlsx)),
            ("tbl_Inmuebles_Compras", lambda: load_inmuebles_compras(sql_conn, inmuebles_xlsx)),
            ("tbl_Inmuebles_MI_Contratos", lambda: load_inmuebles_mi_contratos(sql_conn, inmuebles_xlsx)),
        ]:
            print(f"Cargando {label}...", end=" ")
            sys.stdout.flush()
            try:
                n = fn()
                print(f"OK ({n} filas)")
            except Exception as ex:
                sql_conn.rollback()
                print(f"ERROR: {ex}")
    finally:
        access_conn.close()
        sql_conn.close()

    print("Proceso terminado.")


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--access-path", default=DEFAULT_ACCESS_PATH)
    parser.add_argument("--input-dir", default=DEFAULT_INPUT_DIR)
    parser.add_argument("--sql-server")
    parser.add_argument("--sql-database", default="poc_colombia")
    parser.add_argument("--sql-user")
    parser.add_argument("--check-schema-only", action="store_true")
    args = parser.parse_args()

    if not os.path.isfile(args.access_path):
        print(f"No se encontro el archivo Access en: {args.access_path}")
        sys.exit(1)

    if args.check_schema_only:
        check_schema(args.access_path, args.input_dir)
        return

    if not args.sql_server or not args.sql_user:
        print("Debes indicar --sql-server y --sql-user (o usar --check-schema-only).")
        sys.exit(1)

    password = getpass.getpass(f"Password SQL para {args.sql_user}@{args.sql_server}: ")
    import_data(args.access_path, args.input_dir, args.sql_server, args.sql_database, args.sql_user, password)


if __name__ == "__main__":
    main()
